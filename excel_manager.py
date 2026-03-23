# excel_manager.py
# ─── contacts.xlsx read / write ───────────────────────────────────────────────

import os
import sys
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from config import EXCEL_FILENAME

# ─── Resolve Excel path — always next to the EXE or script ───────────────────

def get_excel_path() -> Path:
    if getattr(sys, "frozen", False):
        base = Path(sys.executable).parent
    else:
        base = Path(__file__).parent
    return base / EXCEL_FILENAME


HEADERS = ["#", "Name", "Phone", "Email", "Source File", "Scanned At"]


# ─── Workbook management ──────────────────────────────────────────────────────

def get_or_create_workbook():
    path = get_excel_path()
    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contacts"

        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            cell.fill      = PatternFill("solid", start_color="1A56DB")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 22

        widths = [5, 28, 18, 34, 36, 20]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    return wb, ws


def append_contact(ws, name: str, phone: str, email: str, source_file: str):
    row    = ws.max_row + 1
    serial = row - 1

    ws.cell(row=row, column=1, value=serial)
    ws.cell(row=row, column=2, value=name)
    ws.cell(row=row, column=3, value=phone)
    ws.cell(row=row, column=4, value=email)
    ws.cell(row=row, column=5, value=source_file)
    ws.cell(row=row, column=6, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

    # Zebra stripe
    if serial % 2 == 0:
        fill = PatternFill("solid", start_color="EEF4FF")
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = fill

    # Consistent font
    for col in range(1, 7):
        ws.cell(row=row, column=col).font = Font(name="Arial", size=10)


def save_workbook(wb):
    wb.save(get_excel_path())


def get_contact_count() -> int:
    path = get_excel_path()
    if not path.exists():
        return 0
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    return max(ws.max_row - 1, 0)


def is_duplicate_email(email: str) -> bool:
    """Check if an email already exists in the sheet."""
    if email == "N/A":
        return False
    path = get_excel_path()
    if not path.exists():
        return False
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3] and str(row[3]).lower() == email.lower():
            return True
    return False


def open_excel():
    """Open the Excel file in the default application."""
    path = get_excel_path()
    if not path.exists():
        return False, "No contacts file yet — scan something first"
    try:
        os.startfile(str(path.resolve()))
        return True, ""
    except Exception as e:
        return False, str(e)
